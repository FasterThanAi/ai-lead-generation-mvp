import csv
import hashlib
import io
import logging
import os
import re
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.core.config import settings
from app.db.models import Product, SourceDocument
from app.services.document_service import extract_text_from_file
from app.utils.time_utils import utc_now

logger = logging.getLogger(__name__)

# Column aliases for fuzzy matching
HEADER_ALIASES = {
    "part_number": {
        "part", "partno", "part no", "part_no", "part-no", "sku", "item",
        "item number", "item_number", "item no", "mfr part", "mfr_part",
        "mfr part number", "mfr_part_no", "mpn", "part #", "model", "model number",
        "partnumber", "product code", "item #"
    },
    "manufacturer": {
        "brand", "mfr", "mfg", "make", "vendor", "supplier", "manufacturer",
        "producer", "brand name", "mfr name"
    },
    "short_description": {
        "description", "desc", "short desc", "short_desc", "short description",
        "name", "title", "product name", "product_name", "item description",
        "product description", "item desc"
    },
    "category": {
        "category", "product category", "type", "product type", "item category",
        "sub category", "subcategory"
    },
    "canonical_name": {
        "canonical name", "canonical_name", "full name", "full_name",
        "standard name", "catalog name"
    },
    "long_description": {
        "long description", "long_description", "details", "extended description",
        "specifications", "summary"
    },
}


def normalize_header(header: str) -> str:
    if not header:
        return ""
    cleaned = str(header).strip().lower()
    cleaned = re.sub(r"[\s_\-]+", " ", cleaned).strip()
    return cleaned


def map_header_to_field(header: str) -> str | None:
    norm = normalize_header(header)
    for field, aliases in HEADER_ALIASES.items():
        if norm == field or norm in aliases:
            return field
        # check without spaces
        norm_no_space = norm.replace(" ", "")
        for alias in aliases:
            if norm_no_space == alias.replace(" ", ""):
                return field
    return None


def sanitize_filename(filename: str) -> str:
    base = os.path.basename(str(filename or "")).strip()
    # remove traversal tokens and path separators
    clean = re.sub(r"[\/\\]+", "", base)
    clean = clean.replace("..", "")
    return clean or "document.bin"


def infer_doc_type(filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower().lstrip(".")
    if ext in {"pdf"}:
        return "pdf"
    if ext in {"docx", "doc"}:
        return "docx"
    if ext in {"png", "jpg", "jpeg", "webp", "gif", "bmp"}:
        return "image"
    if ext in {"csv", "xlsx", "xls"}:
        return "xlsx" if "xls" in ext else "csv"
    if ext in {"txt", "md"}:
        return "text"
    if ext in {"html", "htm"}:
        return "html"
    return ext or "unknown"


def parse_csv_content(file_bytes: bytes) -> list[dict[str, Any]]:
    # try multiple text encodings
    text_content = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text_content = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue

    if text_content is None:
        text_content = file_bytes.decode("utf-8", errors="replace")

    # sniff delimiter
    sample = text_content[:2048]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ","

    reader = csv.reader(io.StringIO(text_content), delimiter=delimiter)
    raw_rows = list(reader)
    if not raw_rows:
        return []

    raw_headers = raw_rows[0]
    column_mapping = {idx: map_header_to_field(h) for idx, h in enumerate(raw_headers)}

    parsed_rows = []
    for row in raw_rows[1:]:
        if not any(str(cell).strip() for cell in row):
            continue
        row_dict = {}
        for idx, cell in enumerate(row):
            field = column_mapping.get(idx)
            if field:
                val = str(cell).strip()
                row_dict[field] = val if val else None
        parsed_rows.append(row_dict)

    return parsed_rows


def parse_xlsx_content(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        if not sheet or sheet.max_row < 1:
            return []

        header_cells = [cell.value for cell in sheet[1]]
        column_mapping = {
            idx: map_header_to_field(str(h) if h is not None else "")
            for idx, h in enumerate(header_cells)
        }

        parsed_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_dict = {}
            for idx, cell in enumerate(row):
                field = column_mapping.get(idx)
                if field:
                    val = str(cell).strip() if cell is not None else ""
                    row_dict[field] = val if val else None
            parsed_rows.append(row_dict)

        return parsed_rows
    except Exception as exc:
        logger.warning("Could not parse XLSX content: %s", exc)
        return []


def parse_product_csv(file_bytes: bytes, filename: str = "") -> dict[str, list[dict[str, Any]]]:
    fn = filename.lower()
    is_xlsx = fn.endswith(".xlsx") or fn.endswith(".xls") or file_bytes.startswith(b"PK\x03\x04")

    if is_xlsx:
        raw_rows = parse_xlsx_content(file_bytes)
    else:
        raw_rows = parse_csv_content(file_bytes)

    valid_rows = []
    rejected_rows = []

    for index, row in enumerate(raw_rows):
        part_number = row.get("part_number")
        if not part_number:
            rejected_rows.append({
                "row_number": index + 2,
                "data": row,
                "reason": "Missing required field: part_number"
            })
        else:
            valid_rows.append({
                "part_number": str(part_number).strip(),
                "manufacturer": row.get("manufacturer"),
                "short_description": row.get("short_description"),
                "category": row.get("category"),
                "canonical_name": row.get("canonical_name"),
                "long_description": row.get("long_description"),
            })

    return {
        "valid": valid_rows,
        "rejected": rejected_rows
    }


def ingest_products(db: Session, catalog_id: int, rows: list[dict[str, Any]]) -> dict[str, int]:
    created = 0
    updated = 0
    rejected = 0

    for row in rows:
        part_number = row.get("part_number")
        if not part_number:
            rejected += 1
            continue

        part_number = str(part_number).strip()
        existing = db.query(Product).filter(
            Product.catalog_id == catalog_id,
            Product.part_number == part_number
        ).first()

        if existing:
            if row.get("manufacturer") is not None:
                existing.manufacturer = row.get("manufacturer")
            if row.get("short_description") is not None:
                existing.short_description = row.get("short_description")
            if row.get("category") is not None:
                existing.category = row.get("category")
            if row.get("canonical_name") is not None:
                existing.canonical_name = row.get("canonical_name")
            if row.get("long_description") is not None:
                existing.long_description = row.get("long_description")
            updated += 1
        else:
            new_product = Product(
                catalog_id=catalog_id,
                part_number=part_number,
                manufacturer=row.get("manufacturer"),
                short_description=row.get("short_description"),
                category=row.get("category"),
                canonical_name=row.get("canonical_name"),
                long_description=row.get("long_description"),
                status="pending",
            )
            db.add(new_product)
            created += 1

    db.commit()

    return {
        "created": created,
        "updated": updated,
        "rejected": rejected,
        "total": len(rows),
    }


def register_document(
    db: Session,
    product_id: int,
    filename: str,
    file_bytes: bytes,
    doc_type: str | None = None
) -> SourceDocument:
    content_hash = hashlib.sha256(file_bytes).hexdigest()

    # Idempotent deduplication per product
    existing = db.query(SourceDocument).filter(
        SourceDocument.product_id == product_id,
        SourceDocument.content_hash == content_hash
    ).first()

    if existing:
        return existing

    clean_name = sanitize_filename(filename)
    inferred_type = doc_type or infer_doc_type(clean_name)
    ext = os.path.splitext(clean_name)[1].lower() or f".{inferred_type}"

    # Save under storage/sources/{product_id}/{content_hash}{ext}
    save_dir = os.path.join(settings.STORAGE_DIR, str(product_id))
    os.makedirs(save_dir, exist_ok=True)
    saved_path = os.path.join(save_dir, f"{content_hash}{ext}")

    with open(saved_path, "wb") as f:
        f.write(file_bytes)

    text_snippet = None
    if inferred_type in {"pdf", "docx", "text", "txt", "md"}:
        try:
            extracted_text = extract_text_from_file(saved_path, inferred_type)
            if extracted_text:
                text_snippet = extracted_text[:2000]
        except Exception as exc:
            logger.warning("Text extraction failed for %s: %s", clean_name, exc)

    source_doc = SourceDocument(
        product_id=product_id,
        filename=clean_name,
        doc_type=inferred_type,
        url=None,
        fetched_at=utc_now(),
        content_hash=content_hash,
        text_snippet=text_snippet,
        page_number=None,
    )
    db.add(source_doc)
    db.commit()
    db.refresh(source_doc)

    return source_doc
